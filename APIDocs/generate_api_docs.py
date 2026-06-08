#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = REPO_ROOT / "sourcecode" / "Royalty%20Calculation.json"
DEFAULT_OUTPUT = Path(__file__).resolve().with_name("Royalty_API_Documentation.xlsx")

TRANSIT_CACHE_BASE = 43
MAP_AS_ARRAY = "^ "
SHORT_IMMUTABLE_TAGS = {"~#iR", "~#iM", "~#iL", "~#iS"}

HEADER_ROW = 3
FIRST_DATA_ROW = 4

COLUMNS = [
    "Workflow",
    "Block",
    "Screen",
    "Action for Operation/Trigger",
    "API ID",
    "API Name",
    "Method",
    "Request",
    "Response",
    "Table",
]

COLUMN_WIDTHS = {
    "A": 24,
    "B": 36,
    "C": 24,
    "D": 40,
    "E": 9,
    "F": 34,
    "G": 12,
    "H": 48,
    "I": 42,
    "J": 46,
}


class ParseError(RuntimeError):
    pass


class TransitTag:
    def __init__(self, tag: str) -> None:
        self.tag = tag


class RollingCache:
    def __init__(self) -> None:
        self._values: list[str] = []

    def encache(self, value: str) -> None:
        self._values.append(value)

    def decode(self, key: str) -> str:
        index = 0
        for char in key[1:]:
            index = (index * TRANSIT_CACHE_BASE) + (ord(char) - 48)
        try:
            return self._values[index]
        except IndexError as exc:
            raise ParseError(f"Retool Transit cache reference {key!r} could not be resolved.") from exc


class TransitDecoder:
    def __init__(self) -> None:
        self.cache = RollingCache()

    def decode(self, node: Any, *, as_map_key: bool = False) -> Any:
        if isinstance(node, str):
            return self._decode_string(node, as_map_key=as_map_key)
        if isinstance(node, list):
            return self._decode_list(node, as_map_key=as_map_key)
        if isinstance(node, dict):
            output: OrderedDict[str, Any] = OrderedDict()
            for key, value in node.items():
                decoded_key = self.decode(key, as_map_key=True)
                decoded_value = self.decode(value, as_map_key=False)
                output[str(decoded_key)] = decoded_value
            return output
        return node

    def _decode_string(self, value: str, *, as_map_key: bool) -> Any:
        if self._is_cache_key(value):
            resolved = self.cache.decode(value)
            return self._parse_string(resolved)

        if self._is_cacheable(value, as_map_key=as_map_key):
            self.cache.encache(value)

        return self._parse_string(value)

    def _decode_list(self, node: list[Any], *, as_map_key: bool) -> Any:
        if not node:
            return []

        first = self.decode(node[0], as_map_key=as_map_key)

        if first == MAP_AS_ARRAY:
            output: OrderedDict[str, Any] = OrderedDict()
            for index in range(1, len(node), 2):
                decoded_key = self.decode(node[index], as_map_key=True)
                decoded_value = self.decode(node[index + 1], as_map_key=False)
                output[str(decoded_key)] = decoded_value
            return output

        if isinstance(first, TransitTag):
            rep = self.decode(node[1], as_map_key=False) if len(node) > 1 else None
            return self._decode_tag(first.tag, rep)

        return [self.decode(item, as_map_key=as_map_key) for item in node]

    @staticmethod
    def _parse_string(value: str) -> Any:
        if not value.startswith("~") or len(value) < 2:
            return value

        marker = value[1]
        if marker == "#":
            return TransitTag(value[2:])
        if marker == "m":
            try:
                milliseconds = int(value[2:])
            except ValueError:
                return value
            return datetime.fromtimestamp(milliseconds / 1000, tz=timezone.utc).isoformat()
        if marker in {"~", "^"}:
            return value[1:]
        return value

    @staticmethod
    def _decode_tag(tag: str, rep: Any) -> Any:
        if tag in {"iM", "iOM"}:
            if not isinstance(rep, list):
                return rep
            output: OrderedDict[str, Any] = OrderedDict()
            for index in range(0, len(rep), 2):
                output[str(rep[index])] = rep[index + 1]
            return output

        if tag in {"iL", "iS"}:
            return list(rep) if isinstance(rep, (list, tuple)) else rep

        if tag == "iR":
            if isinstance(rep, dict):
                return OrderedDict([("__record__", rep.get("n")), ("value", rep.get("v"))])
            return OrderedDict([("__record__", None), ("value", rep)])

        return OrderedDict([("__tag__", tag), ("rep", rep)])

    @staticmethod
    def _is_cache_key(value: str) -> bool:
        return len(value) > 1 and value.startswith("^") and value != MAP_AS_ARRAY

    @staticmethod
    def _is_cacheable(value: str, *, as_map_key: bool) -> bool:
        if len(value) < 4:
            return False
        if as_map_key:
            return True
        return value in SHORT_IMMUTABLE_TAGS or value.startswith("~$") or value.startswith("~:")


def deep_unwrap_records(node: Any) -> Any:
    if isinstance(node, dict):
        normalized = {str(key): deep_unwrap_records(value) for key, value in node.items()}
        if set(normalized.keys()) == {"__record__", "value"}:
            value = normalized["value"]
            if isinstance(value, dict):
                merged = dict(value)
                merged["__record__"] = normalized["__record__"]
                return merged
        return normalized
    if isinstance(node, list):
        return [deep_unwrap_records(item) for item in node]
    return node


def load_retool_plugins(input_path: Path) -> dict[str, dict[str, Any]]:
    try:
        outer_data = json.loads(input_path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise ParseError(f"Retool export JSON was not found: {input_path}") from exc
    except json.JSONDecodeError as exc:
        raise ParseError(f"Retool export JSON is invalid: {input_path}") from exc

    try:
        app_state_raw = outer_data["page"]["data"]["appState"]
    except KeyError as exc:
        raise ParseError("Retool export is missing page.data.appState.") from exc

    if not isinstance(app_state_raw, str):
        raise ParseError("Retool page.data.appState must be a JSON string.")

    decoded = TransitDecoder().decode(json.loads(app_state_raw))
    app_state = deep_unwrap_records(decoded)
    plugins = app_state.get("plugins") if isinstance(app_state, dict) else None
    if not isinstance(plugins, dict):
        raise ParseError("Decoded Retool appState does not contain a plugins map.")

    output: dict[str, dict[str, Any]] = {}
    for key, value in plugins.items():
        if isinstance(value, dict):
            plugin = dict(value)
            plugin.setdefault("id", key)
            if not isinstance(plugin.get("template"), dict):
                plugin["template"] = {}
            output[str(key)] = plugin
    return output


@dataclass(frozen=True)
class ApiRow:
    workflow: str
    block: str
    screen: str
    action: str
    api_name: str
    method: str
    request: str
    response: str
    table: str


def row(
    workflow: str,
    block: str,
    screen: str,
    action: str,
    api_name: str,
    method: str,
    request: str,
    response: str,
    table: str,
) -> ApiRow:
    method = method.upper().strip()
    if method not in {"GET", "POST"}:
        raise ValueError(f"Unsupported method for {api_name}: {method}")
    return ApiRow(workflow, block, screen, action, api_name, method, request, response, table)


def build_api_rows() -> OrderedDict[str, list[ApiRow]]:
    return OrderedDict(
        [
            (
                "Sales",
                [
                    row(
                        "Search sales",
                        "get_sales",
                        "Sales",
                        "User clicks Query Sales Data or uses date filters",
                        "Get Sales Records",
                        "GET",
                        "sales start date, sales end date",
                        "sales rows with item, account, quantity, sales amount, discount, and audit fields",
                        "sales, child_item, account",
                    ),
                    row(
                        "Search sales records",
                        "get_sales_records, get_sales_records_reduce_by_licensor",
                        "Sales",
                        "User opens Sales Records view or changes sales filters",
                        "Get Sales Record Summary",
                        "GET",
                        "start date, end date, licensor, sales account, item filters",
                        "sales record list or grouped sales list by licensor",
                        "parent_item, target_game_titles, sales_accounts, child_item, sales, game_title, contract, licensor",
                    ),
                    row(
                        "Dropdowns",
                        "get_child_item_name, get_sales_accounts",
                        "Sales",
                        "User opens create/update sales modal or a dropdown",
                        "Get Sales Dropdown Options",
                        "GET",
                        "optional search text or selected licensor",
                        "child item options and sales account options",
                        "child_item, sales_accounts",
                    ),
                    row(
                        "Create sales",
                        "add_sales",
                        "Sales",
                        "User clicks Submit in create sales modal",
                        "Create Sales Record",
                        "POST",
                        "child item id, order date, sold quantity, sales amount, manufactured quantity, discount, current user",
                        "created sales record",
                        "sales",
                    ),
                    row(
                        "Update sales",
                        "insert_original_sales_history, insert_new_sales_history, update_sales",
                        "Sales",
                        "User clicks Submit in update sales modal",
                        "Update Sales Record",
                        "POST",
                        "selected sales record, old values, new values, current user",
                        "updated sales record and history result",
                        "sales, sales_history",
                    ),
                    row(
                        "Delete sales",
                        "check_if_record_closed, delete_record_transform, delete_single_record",
                        "Sales",
                        "User confirms Delete Record",
                        "Delete Sales Record",
                        "POST",
                        "selected sales record ids, order date or target month, current user",
                        "deleted count or message that the month is closed",
                        "monthly_royalty_status, sales",
                    ),
                ],
            ),
            (
                "Royalty Calculation",
                [
                    row(
                        "Calculate royalty",
                        "get_all_items",
                        "Royalty Calculation",
                        "User clicks Calculate Monthly Royalty",
                        "Calculate Monthly Royalty",
                        "POST",
                        "start date, end date",
                        "royalty item list with licensor, contract, game title, item, sales, royalty rate, and royalty due",
                        "v_all_transactions",
                    ),
                    row(
                        "Highlight changed price",
                        "get_ids_changed_sales_unit_price_from_last_month",
                        "Royalty Calculation",
                        "After monthly royalty is calculated",
                        "Get Changed Unit Price Items",
                        "GET",
                        "start date",
                        "child item id list where unit price changed from last month",
                        "monthly_royalty",
                    ),
                    row(
                        "Registered royalty lookup",
                        "get_monthly_royalty",
                        "Royalty Calculation",
                        "User searches registered monthly royalty",
                        "Get Registered Monthly Royalty",
                        "GET",
                        "start date, end date",
                        "registered monthly royalty rows",
                        "monthly_royalty",
                    ),
                    row(
                        "Delete registered records",
                        "delete_registered_records",
                        "Royalty Calculation",
                        "User confirms delete registered records",
                        "Delete Registered Royalty Records",
                        "POST",
                        "start date, end date, current user",
                        "deleted monthly royalty and history result",
                        "monthly_royalty, monthly_royalty_history",
                    ),
                    row(
                        "Lookup contract",
                        "read_contract",
                        "Royalty Calculation",
                        "App loads contract reference data",
                        "Get Contract List",
                        "GET",
                        "optional licensor or active flag",
                        "contract and licensor options",
                        "licensor, contract",
                    ),
                ],
            ),
            (
                "Royalty Adjustment",
                [
                    row(
                        "Search adjustments",
                        "get_royalty_adjustment",
                        "Royalty Adjustment",
                        "User clicks Query Royalty Adjustment or uses date filters",
                        "Get Royalty Adjustments",
                        "GET",
                        "adjustment start date, adjustment end date",
                        "royalty adjustment rows",
                        "v_all_royalty_adjustment",
                    ),
                    row(
                        "Adjustment dropdowns",
                        "get_child_item_name, get_licensor_contract_for_create, get_licensor_contract_for_update",
                        "Royalty Adjustment",
                        "User opens create/update adjustment modal or selects child item",
                        "Get Adjustment Dropdown Options",
                        "GET",
                        "selected child item id or search text",
                        "child item, licensor, and contract options",
                        "child_item, licensor, contract, game_title",
                    ),
                    row(
                        "Create adjustment",
                        "add_royalty_adjustment",
                        "Royalty Adjustment",
                        "User clicks Submit in create adjustment modal",
                        "Create New Royalty Adjustment",
                        "POST",
                        "child item id, contract id, adjustment date, royalty adjustment amount, note, current user",
                        "created royalty adjustment record",
                        "royalty_adjustment",
                    ),
                    row(
                        "Update adjustment",
                        "insert_original_royalty_adjustment_history, insert_new_royalty_adjustment_history, update_royalty_adjustment",
                        "Royalty Adjustment",
                        "User clicks Submit in update adjustment modal",
                        "Update Royalty Adjustment",
                        "POST",
                        "selected adjustment id, old values, new values, current user",
                        "updated royalty adjustment and history result",
                        "royalty_adjustment, royalty_adjustment_history",
                    ),
                    row(
                        "Update registered royalty",
                        "update_monthly_royalty",
                        "Royalty Adjustment",
                        "User adjusts a registered monthly royalty record",
                        "Update Monthly Royalty Adjustment",
                        "POST",
                        "monthly royalty id, new royalty amount, adjustment note, current user",
                        "updated monthly royalty record",
                        "monthly_royalty",
                    ),
                    row(
                        "Date defaults",
                        "get_latest_monthly_royalty_dates",
                        "Royalty Adjustment",
                        "Page loads date filters",
                        "Get Latest Monthly Royalty Dates",
                        "GET",
                        "none",
                        "first day and last day for latest monthly royalty period",
                        "v_all_transactions",
                    ),
                ],
            ),
            (
                "Royalty Billing Record",
                [
                    row(
                        "Billing period status",
                        "get_latest_sales_dates, max_closed_year_month, is_status_closed_billing_record",
                        "Royalty Billing Record",
                        "Page loads or date filters change",
                        "Get Billing Period Status",
                        "GET",
                        "optional start date and end date",
                        "latest sales date range and close status",
                        "sales, monthly_royalty_status",
                    ),
                    row(
                        "Billing dropdowns",
                        "read_licensor, get_game_titles_by_lisensor, get_gl_accounts",
                        "Royalty Billing Record",
                        "User changes licensor, game title, or GL account filter",
                        "Get Billing Filter Options",
                        "GET",
                        "selected licensor id and optional search text",
                        "licensor, game title, and GL account options",
                        "licensor, game_title, contract, account",
                    ),
                    row(
                        "MG status",
                        "recouped_contract_list, recouping_contract_list, mgStatusContracts",
                        "Royalty Billing Record",
                        "User clicks Create Royalty Billing Record",
                        "Calculate MG Status For Billing",
                        "POST",
                        "start date, end date, record type",
                        "recouped and recouping contract ids",
                        "v_all_transactions, v_all_royalty_adjustment, contract, sales, licensor",
                    ),
                    row(
                        "Original billing records",
                        "create_open_original_records, create_closed_original_records",
                        "Royalty Billing Record",
                        "User creates original billing records",
                        "Create Original Billing Records",
                        "POST",
                        "start date, end date, record type, selected filters, MG status",
                        "original billing record list",
                        "v_all_transactions, monthly_royalty",
                    ),
                    row(
                        "Changed billing records",
                        "create_changed_records",
                        "Royalty Billing Record",
                        "User creates changed billing records",
                        "Create Changed Billing Records",
                        "POST",
                        "start date, end date, record type, MG status",
                        "changed royalty adjustment billing record list",
                        "v_all_royalty_adjustment",
                    ),
                    row(
                        "Filter billing records",
                        "removed_recouped_credit_records, get_credit_records_filtered_by_licensors, get_credit_records_filtered_by_game_titles, get_credit_records_filtered_by_gl_accounts",
                        "Royalty Billing Record",
                        "User changes licensor, game title, GL account, or All checkboxes",
                        "Filter Billing Records",
                        "POST",
                        "generated billing records, selected licensors, selected game titles, selected GL accounts, all flags",
                        "filtered billing records with return flag",
                        "calculated billing records",
                    ),
                ],
            ),
            (
                "Royalty Report",
                [
                    row(
                        "Report dropdowns",
                        "get_contracts_for_royalty_report, get_game_titles_for_royalty_report",
                        "Royalty Report",
                        "User selects licensor or contract",
                        "Get Report Dropdown Options",
                        "GET",
                        "licensor id, contract id",
                        "contract and game title options",
                        "contract, game_title",
                    ),
                    row(
                        "Royalty due data",
                        "get_monthly_royalty_for_royalty_report",
                        "Royalty Report",
                        "User prepares royalty report",
                        "Get Royalty Report Details",
                        "GET",
                        "licensor id, contract id, report start date, report end date",
                        "monthly royalty rows for report",
                        "v_all_transactions",
                    ),
                    row(
                        "Adjustment data",
                        "get_monthly_adjustments_for_royalty_report",
                        "Royalty Report",
                        "User prepares royalty report",
                        "Get Report Adjustments",
                        "GET",
                        "licensor id, contract id, report start date, report end date",
                        "royalty adjustment rows for report",
                        "v_all_royalty_adjustment",
                    ),
                    row(
                        "Original paid data",
                        "get_original_paid",
                        "Royalty Report",
                        "User prepares royalty report",
                        "Get Original Paid Amount",
                        "GET",
                        "contract id",
                        "original paid amount for selected contract",
                        "contract",
                    ),
                    row(
                        "Accumulated sales",
                        "get_accumlated_units_sold",
                        "Royalty Report",
                        "User prepares royalty report",
                        "Get Accumulated Units Sold",
                        "GET",
                        "licensor id, contract id, end date",
                        "accumulated units sold",
                        "v_all_transactions",
                    ),
                    row(
                        "MG balance report data",
                        "get_mg_balance_for_report, mgBalanceDataForReport",
                        "Royalty Report",
                        "User prepares royalty report",
                        "Calculate Report MG Balance",
                        "POST",
                        "licensor id, contract id, report period",
                        "MG balance data for report",
                        "v_all_transactions, v_all_royalty_adjustment, contract, sales, licensor",
                    ),
                    row(
                        "Export report",
                        "create_report",
                        "Royalty Report",
                        "User clicks Create Report",
                        "Export Royalty Report",
                        "POST",
                        "licensor id, contract id, report start date, report end date",
                        "Excel royalty report file",
                        "v_all_transactions, v_all_royalty_adjustment, contract, game_title",
                    ),
                ],
            ),
            (
                "Close Monthly Royalty",
                [
                    row(
                        "Load close month dates",
                        "get_most_recent_closed_date, get_latest_monthly_royalty_dates",
                        "Close Monthly Royalty",
                        "Page loads close month date filters",
                        "Get Close Month Date Defaults",
                        "GET",
                        "none",
                        "most recent closed date and suggested next period",
                        "monthly_royalty_status, v_all_transactions",
                    ),
                    row(
                        "Check close status",
                        "is_status_record_exists",
                        "Close Monthly Royalty",
                        "User changes close month date range",
                        "Check Monthly Royalty Status",
                        "GET",
                        "close month start date, close month end date",
                        "open or closed status for the month",
                        "monthly_royalty_status",
                    ),
                    row(
                        "Preview close data",
                        "get_royalty_and_adjustment",
                        "Close Monthly Royalty",
                        "User clicks Query Royalty Calculation and Adjustment",
                        "Get Royalty And Adjustments For Close",
                        "GET",
                        "close month start date, close month end date",
                        "royalty and adjustment rows for close review",
                        "v_all_transactions, royalty_adjustment",
                    ),
                    row(
                        "Calculate MG balance",
                        "calc_current_mg_balance, currentMgBalances",
                        "Close Monthly Royalty",
                        "User clicks Calculate MG Balance",
                        "Calculate Current MG Balance",
                        "POST",
                        "close month start date, close month end date",
                        "current MG balance by contract",
                        "v_all_transactions, v_all_royalty_adjustment, contract, sales, licensor",
                    ),
                    row(
                        "Register monthly royalty",
                        "transform_royalty_and_adjustment_table_to_insert_monthly_royalty, insert_monthly_royalty",
                        "Close Monthly Royalty",
                        "User clicks Register Royalty",
                        "Register Monthly Royalty",
                        "POST",
                        "approved royalty and adjustment rows, current user",
                        "inserted monthly royalty result",
                        "monthly_royalty",
                    ),
                    row(
                        "Close month",
                        "close_month, get_max_status_id, insert_monthly_status, insert_original_monthly_status_history, insert_new_monthly_status_history",
                        "Close Monthly Royalty",
                        "User clicks Register and Close Monthly Royalty",
                        "Close Monthly Royalty",
                        "POST",
                        "start date, end date, current user, close flag",
                        "closed monthly status and history result",
                        "monthly_royalty_status, monthly_royalty_status_history",
                    ),
                ],
            ),
            (
                "Master Data - Utilities",
                [
                    row(
                        "Lookup licensors",
                        "read_licensor",
                        "Master Data / Utilities",
                        "Any page needs licensor dropdown options",
                        "Get Licensors",
                        "GET",
                        "optional search text",
                        "licensor list",
                        "licensor",
                    ),
                    row(
                        "Lookup game titles",
                        "get_game_titles",
                        "Master Data / Utilities",
                        "Any page needs game title dropdown options",
                        "Get Game Titles",
                        "GET",
                        "optional licensor id, contract id, or search text",
                        "game title list",
                        "game_title",
                    ),
                    row(
                        "Lookup GL accounts",
                        "get_gl_accounts",
                        "Master Data / Utilities",
                        "Any page needs GL account dropdown options",
                        "Get GL Accounts",
                        "GET",
                        "optional account type or search text",
                        "GL account list",
                        "account",
                    ),
                    row(
                        "Lookup active contracts",
                        "get_active_contracts",
                        "Master Data / Utilities",
                        "MG balance calculation needs active contracts",
                        "Get Active Contracts",
                        "GET",
                        "optional licensor id",
                        "active contract list",
                        "contract, licensor",
                    ),
                    row(
                        "Utility royalty due open",
                        "get_royalty_due_from_open_month",
                        "Master Data / Utilities",
                        "MG balance calculation reads open month royalty due",
                        "Get Royalty Due From Open Month",
                        "GET",
                        "start date, end date",
                        "royalty due from open transactions",
                        "v_all_transactions",
                    ),
                    row(
                        "Utility royalty due closed",
                        "get_royalty_due_from_closed_month",
                        "Master Data / Utilities",
                        "MG balance calculation reads closed month royalty due",
                        "Get Royalty Due From Closed Month",
                        "GET",
                        "start date, end date",
                        "royalty due from registered monthly royalty",
                        "monthly_royalty",
                    ),
                    row(
                        "Utility adjustment for MG",
                        "get_royalty_adjustment_for_calculation_mg",
                        "Master Data / Utilities",
                        "MG balance calculation reads royalty adjustments",
                        "Get Royalty Adjustment For MG Calculation",
                        "GET",
                        "start date, end date",
                        "royalty adjustments for MG calculation",
                        "v_all_royalty_adjustment",
                    ),
                    row(
                        "Shared MG calculation",
                        "calculateMGBalance",
                        "Master Data / Utilities",
                        "Billing, close month, or report needs MG balance",
                        "Calculate MG Balance",
                        "POST",
                        "start date, end date, optional max closed month",
                        "prior and current MG balance by licensor and contract, with recouped status",
                        "monthly_royalty, v_all_transactions, v_all_royalty_adjustment, contract, sales, licensor",
                    ),
                    row(
                        "Transfer lookup",
                        "get_contracts_by_transfer_from_licensor, get_contracts_by_transfer_to_licensor, get_transfer_contract_history",
                        "Master Data / Utilities",
                        "User selects transfer from/to licensor or opens transfer history",
                        "Get Transfer Contract Data",
                        "GET",
                        "from licensor id, to licensor id, optional history licensor id",
                        "transferable contract options and transfer history",
                        "active_contract, contract, licensor",
                    ),
                    row(
                        "Transfer contracts",
                        "transfer_licensors, insert_transfer_contract, update_active_contract",
                        "Master Data / Utilities",
                        "User submits transfer licensors form",
                        "Transfer Licensor Contracts",
                        "POST",
                        "selected contract ids, transfer date, new licensor id, new version, current user",
                        "inserted contract versions and updated active contract result",
                        "contract, active_contract",
                    ),
                    row(
                        "Delete transfer",
                        "delete_transfered_contract, update_transfered_contract_for_delete",
                        "Master Data / Utilities",
                        "User deletes a transferred contract history row",
                        "Delete Transferred Contract",
                        "POST",
                        "selected transferred contract row, current user",
                        "deleted or reverted transferred contract result",
                        "contract, active_contract",
                    ),
                ],
            ),
        ]
    )


def extract_block_ids(block: str) -> list[str]:
    ids: list[str] = []
    for part in block.split(","):
        candidate = part.strip()
        if candidate:
            ids.append(candidate)
    return ids


def validate_blocks(rows_by_sheet: OrderedDict[str, list[ApiRow]], plugins: dict[str, dict[str, Any]]) -> dict[str, list[str]]:
    missing: dict[str, list[str]] = {}
    for sheet, rows in rows_by_sheet.items():
        for api_row in rows:
            for block_id in extract_block_ids(api_row.block):
                if block_id not in plugins:
                    missing.setdefault(sheet, []).append(block_id)
    return missing


def create_workbook(rows_by_sheet: OrderedDict[str, list[ApiRow]], output_path: Path) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    wb.properties.title = "Royalty API Documentation"
    wb.properties.subject = "Retool migration API workbook"
    wb.properties.creator = "Codex"

    for sheet_name, rows in rows_by_sheet.items():
        ws = wb.create_sheet(title=safe_sheet_name(sheet_name))
        write_sheet(ws, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "-", name)
    return cleaned[:31]


def write_sheet(ws: Any, rows: list[ApiRow]) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    ws.cell(row=1, column=1).value = (
        "Migration note: Current architecture is Retool -> Database. "
        "Target architecture is Retool -> Backend API -> Database. Retool should only call APIs."
    )
    ws.cell(row=2, column=7).value = "GET means only returning data from DB to show. POST means interacting with DB."
    ws.cell(row=2, column=8).value = "What data Retool sends to the API. For GET, include filter fields."
    ws.cell(row=2, column=9).value = "What API sends back to Retool."
    ws.cell(row=2, column=10).value = "What tables are going to be looked at or changed."

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.value = header

    for api_id, api_row in enumerate(rows, start=1):
        row_idx = FIRST_DATA_ROW + api_id - 1
        values = [
            api_row.workflow,
            api_row.block,
            api_row.screen,
            api_row.action,
            api_id,
            api_row.api_name,
            api_row.method,
            api_row.request,
            api_row.response,
            api_row.table,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    style_sheet(ws, len(rows))


def style_sheet(ws: Any, data_row_count: int) -> None:
    black_fill = PatternFill("solid", fgColor="000000")
    white_font = Font(color="FFFFFF", bold=True)
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    instruction_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    ws.cell(row=1, column=1).fill = note_fill
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 36

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = instruction_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 48

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx)
        cell.fill = black_fill
        cell.font = white_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = border
    ws.row_dimensions[HEADER_ROW].height = 28

    max_row = FIRST_DATA_ROW + data_row_count - 1
    for row_idx in range(FIRST_DATA_ROW, max_row + 1):
        ws.row_dimensions[row_idx].height = 54
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            horizontal = "center" if col_idx in {5, 7} else "left"
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal=horizontal)
            cell.border = border

    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(COLUMNS))}{max_row}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Royalty Retool API documentation workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Retool exported JSON file.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output .xlsx workbook path.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    plugins = load_retool_plugins(args.input)
    rows_by_sheet = build_api_rows()
    missing = validate_blocks(rows_by_sheet, plugins)
    create_workbook(rows_by_sheet, args.output)

    print(f"Wrote workbook: {args.output}")
    if missing:
        print("Source blocks not found in export:")
        for sheet, block_ids in missing.items():
            unique_ids = sorted(set(block_ids))
            print(f"  {sheet}: {', '.join(unique_ids)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
